#!/usr/bin/env python3
"""
Job Hunter - AI-powered job hunting agent
Scrapes job boards, scores jobs against resume using AI, outputs Excel with rich formatting
"""

import asyncio
import json
import logging
import os
import sys
import re
import random
import argparse
from dataclasses import dataclass, asdict, field
from datetime import datetime
from typing import Optional, List, Dict, Any
from pathlib import Path
from collections import Counter
import base64

import aiohttp
from aiohttp import TCPConnector
import asyncio
from asyncio import Semaphore

import pdfplumber
from docx import Document
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from fake_useragent import UserAgent

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from tqdm import tqdm
from dotenv import load_dotenv

try:
    from playwright.async_api import async_playwright, Browser
except ImportError:
    Browser = None

try:
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import Mail, Email, To, Content
    HAS_SENDGRID = True
except ImportError:
    HAS_SENDGRID = False

# ============================================================================
# LOGGING SETUP
# ============================================================================

def setup_logging(log_level: str = "INFO"):
    """Setup logging to file and stdout"""
    log_dir = Path(".")
    log_file = log_dir / "job_hunter.log"
    
    logging.basicConfig(
        level=getattr(logging, log_level.upper()),
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

logger = None

# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class ResumeProfile:
    """Parsed resume profile"""
    full_text: str
    skills: List[str]
    experience_years: int
    education: List[str]
    job_titles: List[str]
    certifications: List[str]
    tech_keywords: List[str]

@dataclass
class RawJob:
    """Raw job from scraper"""
    title: str
    company: str
    description: str
    url: str
    source_board: str

@dataclass
class JobResult:
    """Fully analyzed job"""
    # IDENTIFICATION
    job_title: str
    company: str
    location: str
    remote_type: str
    employment_type: str
    experience_level: str
    industry: str
    company_size: str
    
    # COMPENSATION & LOGISTICS
    salary_min: Optional[int]
    salary_max: Optional[int]
    salary_currency: Optional[str]
    salary_display: str
    visa_sponsorship: str
    posted_date: Optional[str]
    application_deadline: Optional[str]
    
    # CONTENT
    description_summary: str
    skills_required: List[str]
    responsibilities: List[str]
    nice_to_have: List[str]
    
    # ATS SCORING
    ats_score: int
    ats_breakdown: Dict[str, int]
    match_verdict: str
    keywords_matched: List[str]
    keywords_missing: List[str]
    skill_gaps: List[str]
    
    # ACTIONABLE ADVICE
    resume_tips: List[str]
    cover_letter_hint: str
    interview_topics: List[str]
    
    # CONTACT & APPLICATION
    contact_email: Optional[str]
    contact_phone: Optional[str]
    apply_url: str
    
    # META
    source_board: str
    ai_provider_used: str
    run_timestamp: str

# ============================================================================
# AI CLIENT - PROVIDER ADAPTER
# ============================================================================

class AIClient:
    """Unified interface for 5 AI providers"""
    
    def __init__(self, provider: str, api_key: str):
        self.provider = provider
        self.api_key = api_key
        self.session = None
        self.ua = UserAgent()
        
        # Provider configurations
        self.config = {
            "claude": {
                "base_url": "https://api.anthropic.com/v1/messages",
                "model": "claude-sonnet-4-20250514"
            },
            "gemini": {
                "base_url": "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-pro:generateContent",
                "model": "gemini-1.5-pro"
            },
            "groq": {
                "base_url": "https://api.groq.com/openai/v1/chat/completions",
                "model": "llama-3.3-70b-versatile"
            },
            "deepseek": {
                "base_url": "https://api.deepseek.com/v1/chat/completions",
                "model": "deepseek-chat"
            },
            "nvidia": {
                "base_url": "https://integrate.api.nvidia.com/v1/chat/completions",
                "model": "meta/llama-3.1-70b-instruct"
            }
        }
    
    async def __aenter__(self):
        connector = TCPConnector(limit_per_host=10, limit=100)
        self.session = aiohttp.ClientSession(connector=connector)
        return self
    
    async def __aexit__(self, exc_type, exc, tb):
        if self.session:
            await self.session.close()
    
    async def ping(self) -> bool:
        """Health check - minimal test call"""
        try:
            if self.provider == "claude":
                return await self._ping_claude()
            elif self.provider == "gemini":
                return await self._ping_gemini()
            elif self.provider in ["groq", "deepseek", "nvidia"]:
                return await self._ping_openai_compat()
            return False
        except Exception as e:
            logger.warning(f"Ping failed for {self.provider}: {e}")
            return False
    
    async def _ping_claude(self) -> bool:
        """Test Claude API"""
        headers = {
            "x-api-key": self.api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json"
        }
        body = {
            "model": self.config["claude"]["model"],
            "max_tokens": 100,
            "messages": [{"role": "user", "content": "Say OK"}]
        }
        async with self.session.post(self.config["claude"]["base_url"], 
                                     headers=headers, json=body, timeout=aiohttp.ClientTimeout(total=10)) as resp:
            return resp.status == 200
    
    async def _ping_gemini(self) -> bool:
        """Test Gemini API"""
        url = f"{self.config['gemini']['base_url']}?key={self.api_key}"
        body = {
            "contents": [{"parts": [{"text": "Say OK"}]}]
        }
        async with self.session.post(url, json=body, timeout=aiohttp.ClientTimeout(total=10)) as resp:
            return resp.status == 200
    
    async def _ping_openai_compat(self) -> bool:
        """Test OpenAI-compatible API (Groq, DeepSeek, Nvidia)"""
        config = self.config[self.provider]
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "content-type": "application/json"
        }
        body = {
            "model": config["model"],
            "messages": [{"role": "user", "content": "Say OK"}],
            "max_tokens": 100
        }
        async with self.session.post(config["base_url"], headers=headers, json=body, 
                                     timeout=aiohttp.ClientTimeout(total=10)) as resp:
            return resp.status == 200
    
    async def complete(self, prompt: str, system: str = "") -> str:
        """Get completion from active provider with exponential backoff"""
        max_retries = 3
        base_delay = 1.0
        max_delay = 16.0
        
        for attempt in range(max_retries):
            try:
                if self.provider == "claude":
                    return await self._complete_claude(prompt, system)
                elif self.provider == "gemini":
                    return await self._complete_gemini(prompt, system)
                elif self.provider in ["groq", "deepseek", "nvidia"]:
                    return await self._complete_openai_compat(prompt, system)
            except Exception as e:
                if attempt < max_retries - 1:
                    # Exponential backoff with jitter
                    delay = min(base_delay * (2 ** attempt) + random.uniform(0, 1), max_delay)
                    logger.warning(f"Provider {self.provider} attempt {attempt + 1} failed: {e}. Retrying in {delay:.1f}s...")
                    await asyncio.sleep(delay)
                else:
                    logger.error(f"Provider {self.provider} failed after {max_retries} attempts: {e}")
                    raise
        
        raise RuntimeError(f"Failed to get completion from {self.provider}")
    
    async def _complete_claude(self, prompt: str, system: str = "") -> str:
        """Claude completion"""
        headers = {
            "x-api-key": self.api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json"
        }
        body = {
            "model": self.config["claude"]["model"],
            "max_tokens": 2048,
            "system": system or "You are a helpful assistant.",
            "messages": [{"role": "user", "content": prompt}]
        }
        async with self.session.post(self.config["claude"]["base_url"], 
                                     headers=headers, json=body, timeout=aiohttp.ClientTimeout(total=60)) as resp:
            if resp.status != 200:
                raise RuntimeError(f"Claude API error: {resp.status} - {await resp.text()}")
            data = await resp.json()
            return data["content"][0]["text"]
    
    async def _complete_gemini(self, prompt: str, system: str = "") -> str:
        """Gemini completion"""
        url = f"{self.config['gemini']['base_url']}?key={self.api_key}"
        body = {
            "contents": [{"parts": [{"text": prompt}]}],
            "systemInstruction": {"parts": [{"text": system or "You are a helpful assistant."}]}
        }
        async with self.session.post(url, json=body, timeout=aiohttp.ClientTimeout(total=60)) as resp:
            if resp.status != 200:
                raise RuntimeError(f"Gemini API error: {resp.status} - {await resp.text()}")
            data = await resp.json()
            return data["candidates"][0]["content"]["parts"][0]["text"]
    
    async def _complete_openai_compat(self, prompt: str, system: str = "") -> str:
        """OpenAI-compatible completion (Groq, DeepSeek, Nvidia)"""
        config = self.config[self.provider]
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "content-type": "application/json"
        }
        messages = []
        if system:
            messages.append({"role": "system", "content": system})
        messages.append({"role": "user", "content": prompt})
        
        body = {
            "model": config["model"],
            "messages": messages,
            "max_tokens": 2048
        }
        async with self.session.post(config["base_url"], headers=headers, json=body, 
                                     timeout=aiohttp.ClientTimeout(total=60)) as resp:
            if resp.status != 200:
                raise RuntimeError(f"{self.provider} API error: {resp.status} - {await resp.text()}")
            data = await resp.json()
            return data["choices"][0]["message"]["content"]

# ============================================================================
# PROVIDER DETECTION & INITIALIZATION
# ============================================================================

async def detect_active_provider(force_provider: Optional[str] = None) -> tuple[AIClient, str]:
    """
    Detect and initialize the active AI provider
    Priority: Claude → Gemini → Groq → DeepSeek → Nvidia
    """
    providers = [
        ("claude", "ANTHROPIC_API_KEY"),
        ("gemini", "GEMINI_API_KEY"),
        ("groq", "GROQ_API_KEY"),
        ("deepseek", "DEEPSEEK_API_KEY"),
        ("nvidia", "NVIDIA_API_KEY"),
    ]
    
    if force_provider:
        for provider_name, env_key in providers:
            if provider_name.lower() == force_provider.lower():
                api_key = os.getenv(env_key, "").strip()
                if not api_key:
                    logger.error(f"Forced provider {force_provider} selected but {env_key} is empty")
                    sys.exit(1)
                async with AIClient(provider_name, api_key) as client:
                    if not await client.ping():
                        logger.error(f"Provider {force_provider} ping failed")
                        sys.exit(1)
                logger.info(f"Using forced provider: {force_provider}")
                return AIClient(provider_name, api_key), provider_name
        logger.error(f"Unknown provider: {force_provider}")
        sys.exit(1)
    
    for provider_name, env_key in providers:
        api_key = os.getenv(env_key, "").strip()
        if api_key:
            logger.info(f"Testing provider {provider_name}...")
            async with AIClient(provider_name, api_key) as client:
                if await client.ping():
                    logger.info(f"✓ Active provider: {provider_name}")
                    return AIClient(provider_name, api_key), provider_name
                else:
                    logger.warning(f"✗ Provider {provider_name} ping failed, trying next...")
    
    logger.error("No AI provider keys found. Please set at least one of:")
    for _, env_key in providers:
        logger.error(f"  - {env_key}")
    sys.exit(1)

# ============================================================================
# RESUME PARSER
# ============================================================================

def parse_resume_file(file_path: str) -> str:
    """Parse resume PDF or DOCX and return full text"""
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"Resume file not found: {file_path}")
    
    if file_path.suffix.lower() == ".pdf":
        return parse_pdf(file_path)
    elif file_path.suffix.lower() == ".docx":
        return parse_docx(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_path.suffix}. Use PDF or DOCX.")

def parse_pdf(file_path: Path) -> str:
    """Extract text from PDF"""
    text = []
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
    except Exception as e:
        logger.error(f"Error parsing PDF: {e}")
        return ""
    return "\n".join(text)

def parse_docx(file_path: Path) -> str:
    """Extract text from DOCX"""
    text = []
    try:
        doc = Document(file_path)
        for para in doc.paragraphs:
            if para.text.strip():
                text.append(para.text)
    except Exception as e:
        logger.error(f"Error parsing DOCX: {e}")
        return ""
    return "\n".join(text)

async def extract_resume_profile(resume_text: str, ai_client: AIClient) -> ResumeProfile:
    """
    Use AI to extract structured resume profile
    Fall back to regex if AI response is not valid JSON
    """
    prompt = f"""Analyze this resume and extract structured information. Return ONLY a valid JSON object (no markdown, no preamble) with these exact fields:
{{
    "skills": ["skill1", "skill2", ...],
    "experience_years": <integer>,
    "education": ["degree1", "degree2", ...],
    "job_titles": ["title1", "title2", ...],
    "certifications": ["cert1", "cert2", ...],
    "tech_keywords": ["keyword1", "keyword2", ...]
}}

Resume text:
{resume_text[:2000]}
"""
    
    try:
        response = await ai_client.complete(prompt)
        # Try to parse JSON
        data = json.loads(response)
        logger.info("Resume profile extracted via AI")
        return ResumeProfile(
            full_text=resume_text,
            skills=data.get("skills", []),
            experience_years=data.get("experience_years", 0),
            education=data.get("education", []),
            job_titles=data.get("job_titles", []),
            certifications=data.get("certifications", []),
            tech_keywords=data.get("tech_keywords", [])
        )
    except (json.JSONDecodeError, KeyError) as e:
        logger.warning(f"AI extraction failed, using regex fallback: {e}")
        return extract_resume_profile_regex(resume_text)

def extract_resume_profile_regex(resume_text: str) -> ResumeProfile:
    """Fallback: extract resume profile using regex"""
    text_lower = resume_text.lower()
    
    # Extract experience years
    exp_match = re.search(r'(\d+)\s*(?:years?|yrs?)\s*(?:of\s*)?(?:experience|exp)', text_lower)
    experience_years = int(exp_match.group(1)) if exp_match else 0
    
    # Common tech keywords
    tech_keywords = [
        "python", "javascript", "java", "c++", "c#", "go", "rust", "kotlin",
        "react", "angular", "vue", "node", "django", "flask", "fastapi",
        "aws", "gcp", "azure", "docker", "kubernetes", "terraform",
        "sql", "postgresql", "mongodb", "redis", "elasticsearch",
        "git", "jenkins", "gitlab", "github", "ci/cd"
    ]
    
    matched_keywords = [kw for kw in tech_keywords if kw in text_lower]
    
    # Extract education (simple pattern)
    education = []
    for pattern in [r"(?:B\.?S\.?|B\.?A\.?|M\.?S\.?|M\.?B\.?A\.?|Ph\.?D\.?) (?:in\s+)?([^,\n]+)",
                    r"(?:Bachelor|Master|Doctor) (?:of\s+)?(?:Science|Arts|Engineering) (?:in\s+)?([^,\n]+)"]:
        matches = re.findall(pattern, resume_text, re.IGNORECASE)
        education.extend(matches)
    
    return ResumeProfile(
        full_text=resume_text,
        skills=[],
        experience_years=experience_years,
        education=education[:5],
        job_titles=[],
        certifications=[],
        tech_keywords=matched_keywords
    )

# ============================================================================
# JOB SCRAPERS
# ============================================================================

class JobScraper:
    """Base job scraper using Playwright"""
    
    def __init__(self):
        self.ua = UserAgent()
    
    async def scrape_jobs(self, search_query: str, location: str, limit: int = 20) -> List[RawJob]:
        """Override in subclasses"""
        return []

class LinkedInScraper(JobScraper):
    """LinkedIn Jobs scraper"""
    
    async def scrape_jobs(self, search_query: str, location: str, limit: int = 20) -> List[RawJob]:
        """Scrape LinkedIn jobs"""
        jobs = []
        try:
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                page = await browser.new_page()
                
                # LinkedIn search URL
                search_url = f"https://www.linkedin.com/jobs/search/?keywords={search_query}&location={location}"
                await page.goto(search_url, wait_until="networkidle", timeout=30000)
                
                # Wait for job listings
                await page.wait_for_selector('[data-job-id]', timeout=10000)
                
                # Extract job listings
                jobs_html = await page.locator('[data-job-id]').all()
                
                for job_elem in jobs_html[:limit]:
                    try:
                        job_id = await job_elem.get_attribute('data-job-id')
                        job_title = await job_elem.locator('h3').text_content()
                        company = await job_elem.locator('[data-test="company-name"]').text_content()
                        url = f"https://www.linkedin.com/jobs/view/{job_id}/"
                        
                        # Click to get description
                        await job_elem.click()
                        await page.wait_for_timeout(500)
                        description = await page.locator('[id*="description"]').text_content()
                        
                        if job_title and company and description:
                            jobs.append(RawJob(
                                title=job_title.strip(),
                                company=company.strip(),
                                description=description.strip()[:1000],
                                url=url,
                                source_board="LinkedIn"
                            ))
                    except Exception as e:
                        logger.debug(f"Error extracting LinkedIn job: {e}")
                        continue
                
                await browser.close()
                logger.info(f"LinkedIn: found {len(jobs)} jobs")
        except Exception as e:
            logger.warning(f"LinkedIn scraper failed: {e}")
        
        return jobs

class IndeedScraper(JobScraper):
    """Indeed scraper using BeautifulSoup"""
    
    async def scrape_jobs(self, search_query: str, location: str, limit: int = 20) -> List[RawJob]:
        """Scrape Indeed jobs"""
        jobs = []
        try:
            async with aiohttp.ClientSession() as session:
                url = f"https://www.indeed.com/jobs?q={search_query}&l={location}&limit={limit}"
                headers = {"User-Agent": self.ua.random}
                
                await asyncio.sleep(random.uniform(1, 3))
                async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=30)) as resp:
                    if resp.status == 200:
                        html = await resp.text()
                        soup = BeautifulSoup(html, 'html.parser')
                        
                        for job_card in soup.find_all('div', class_='job_seen_beacon')[:limit]:
                            try:
                                title_elem = job_card.find('h2', class_='jobTitle')
                                company_elem = job_card.find('span', class_='companyName')
                                desc_elem = job_card.find('div', class_='job-snippet')
                                link_elem = job_card.find('a', class_='jcs-DesktopSerpJob')
                                
                                if title_elem and company_elem and link_elem:
                                    title = title_elem.get_text(strip=True)
                                    company = company_elem.get_text(strip=True)
                                    description = desc_elem.get_text(strip=True) if desc_elem else ""
                                    url = f"https://www.indeed.com{link_elem.get('href', '')}"
                                    
                                    jobs.append(RawJob(
                                        title=title,
                                        company=company,
                                        description=description[:500],
                                        url=url,
                                        source_board="Indeed"
                                    ))
                            except Exception as e:
                                logger.debug(f"Error extracting Indeed job: {e}")
                                continue
                        
                        logger.info(f"Indeed: found {len(jobs)} jobs")
        except Exception as e:
            logger.warning(f"Indeed scraper failed: {e}")
        
        return jobs

class GlassdoorScraper(JobScraper):
    """Glassdoor scraper"""
    
    async def scrape_jobs(self, search_query: str, location: str, limit: int = 20) -> List[RawJob]:
        """Scrape Glassdoor jobs"""
        jobs = []
        try:
            async with aiohttp.ClientSession() as session:
                url = f"https://www.glassdoor.com/Job/jobs.htm?keyword={search_query}&location={location}"
                headers = {"User-Agent": self.ua.random}
                
                await asyncio.sleep(random.uniform(1, 3))
                async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=30)) as resp:
                    if resp.status == 200:
                        html = await resp.text()
                        soup = BeautifulSoup(html, 'html.parser')
                        
                        for job_card in soup.find_all('li', class_='jl')[:limit]:
                            try:
                                title_elem = job_card.find('a', class_='job_link')
                                company_elem = job_card.find('div', class_='employer')
                                
                                if title_elem and company_elem:
                                    title = title_elem.get_text(strip=True)
                                    company = company_elem.get_text(strip=True)
                                    url = f"https://www.glassdoor.com{title_elem.get('href', '')}"
                                    
                                    jobs.append(RawJob(
                                        title=title,
                                        company=company,
                                        description="",
                                        url=url,
                                        source_board="Glassdoor"
                                    ))
                            except Exception as e:
                                logger.debug(f"Error extracting Glassdoor job: {e}")
                                continue
                        
                        logger.info(f"Glassdoor: found {len(jobs)} jobs")
        except Exception as e:
            logger.warning(f"Glassdoor scraper failed: {e}")
        
        return jobs

class NaukriScraper(JobScraper):
    """Naukri scraper"""
    
    async def scrape_jobs(self, search_query: str, location: str, limit: int = 20) -> List[RawJob]:
        """Scrape Naukri jobs"""
        jobs = []
        try:
            async with aiohttp.ClientSession() as session:
                url = f"https://www.naukri.com/search?keyword={search_query}&location={location}"
                headers = {"User-Agent": self.ua.random}
                
                await asyncio.sleep(random.uniform(1, 3))
                async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=30)) as resp:
                    if resp.status == 200:
                        html = await resp.text()
                        soup = BeautifulSoup(html, 'html.parser')
                        
                        for job_card in soup.find_all('article', class_='jobTuple')[:limit]:
                            try:
                                title_elem = job_card.find('a', class_='jobTitle')
                                company_elem = job_card.find('a', class_='subTitle')
                                
                                if title_elem and company_elem:
                                    title = title_elem.get_text(strip=True)
                                    company = company_elem.get_text(strip=True)
                                    url = title_elem.get('href', '')
                                    
                                    jobs.append(RawJob(
                                        title=title,
                                        company=company,
                                        description="",
                                        url=url,
                                        source_board="Naukri"
                                    ))
                            except Exception as e:
                                logger.debug(f"Error extracting Naukri job: {e}")
                                continue
                        
                        logger.info(f"Naukri: found {len(jobs)} jobs")
        except Exception as e:
            logger.warning(f"Naukri scraper failed: {e}")
        
        return jobs

class WellfoundScraper(JobScraper):
    """Wellfound (AngelList) scraper"""
    
    async def scrape_jobs(self, search_query: str, location: str, limit: int = 20) -> List[RawJob]:
        """Scrape Wellfound jobs"""
        jobs = []
        try:
            async with aiohttp.ClientSession() as session:
                url = f"https://wellfound.com/jobs?keywords={search_query}&locations={location}"
                headers = {"User-Agent": self.ua.random}
                
                await asyncio.sleep(random.uniform(1, 3))
                async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=30)) as resp:
                    if resp.status == 200:
                        html = await resp.text()
                        soup = BeautifulSoup(html, 'html.parser')
                        
                        for job_card in soup.find_all('div', class_='job-card')[:limit]:
                            try:
                                title_elem = job_card.find('h2', class_='job-title')
                                company_elem = job_card.find('div', class_='company-name')
                                
                                if title_elem and company_elem:
                                    title = title_elem.get_text(strip=True)
                                    company = company_elem.get_text(strip=True)
                                    
                                    jobs.append(RawJob(
                                        title=title,
                                        company=company,
                                        description="",
                                        url="https://wellfound.com",
                                        source_board="Wellfound"
                                    ))
                            except Exception as e:
                                logger.debug(f"Error extracting Wellfound job: {e}")
                                continue
                        
                        logger.info(f"Wellfound: found {len(jobs)} jobs")
        except Exception as e:
            logger.warning(f"Wellfound scraper failed: {e}")
        
        return jobs

async def scrape_all_jobs(search_query: str, location: str, boards: List[str], limit: int = 20) -> List[RawJob]:
    """Scrape jobs from all selected boards concurrently"""
    scrapers = {
        "linkedin": LinkedInScraper(),
        "indeed": IndeedScraper(),
        "glassdoor": GlassdoorScraper(),
        "naukri": NaukriScraper(),
        "wellfound": WellfoundScraper()
    }
    
    tasks = []
    for board in boards:
        if board.lower() in scrapers:
            tasks.append(scrapers[board.lower()].scrape_jobs(search_query, location, limit))
    
    results = await asyncio.gather(*tasks, return_exceptions=True)
    
    all_jobs = []
    for result in results:
        if isinstance(result, Exception):
            logger.warning(f"Scraper error: {result}")
        elif result:
            all_jobs.extend(result)
    
    # Deduplicate by (company.lower() + title.lower())
    seen = set()
    deduplicated = []
    for job in all_jobs:
        key = (job.company.lower(), job.title.lower())
        if key not in seen:
            seen.add(key)
            deduplicated.append(job)
    
    logger.info(f"Total jobs scraped: {len(all_jobs)}, after dedup: {len(deduplicated)}")
    return deduplicated

# ============================================================================
# JOB ANALYSIS WITH AI
# ============================================================================

async def analyze_job(job: RawJob, resume: ResumeProfile, ai_client: AIClient, provider_name: str) -> Optional[JobResult]:
    """
    Analyze a single job using AI
    Returns JobResult or None if analysis fails
    """
    prompt = f"""Analyze this job posting and the provided resume. Return ONLY a valid JSON object with these fields:
{{
    "job_title": "normalized job title",
    "location": "city, country",
    "remote_type": "Remote|Hybrid|On-site|Unknown",
    "employment_type": "Full-time|Part-time|Contract|Freelance|Internship",
    "experience_level": "Internship|Junior|Mid|Senior|Lead|Executive",
    "industry": "industry/domain",
    "company_size": "1-10|11-50|51-200|201-500|501-1000|1000+|Unknown",
    "salary_min": <int or null>,
    "salary_max": <int or null>,
    "salary_currency": "USD|INR|GBP|EUR|...",
    "salary_display": "human readable string",
    "visa_sponsorship": "Yes|No|Unknown",
    "posted_date": "YYYY-MM-DD or null",
    "application_deadline": "YYYY-MM-DD or null",
    "description_summary": "3-sentence plain English summary",
    "skills_required": ["skill1", "skill2", ...],
    "responsibilities": ["resp1", "resp2", "resp3", "resp4", "resp5"],
    "nice_to_have": ["skill1", "skill2", ...],
    "ats_score": <0-100>,
    "keywords_matched": ["matched1", "matched2", ...],
    "keywords_missing": ["gap1", "gap2", ...],
    "skill_gaps": ["skill to learn 1", "skill to learn 2", ...],
    "resume_tips": ["tip1", "tip2", "tip3"],
    "cover_letter_hint": "2-3 key points to emphasise",
    "interview_topics": ["topic1", "topic2", "topic3"],
    "contact_email": "email or null",
    "contact_phone": "phone or null"
}}

Job Title: {job.title}
Company: {job.company}
Job Description: {job.description}
Job URL: {job.url}

Resume Skills: {', '.join(resume.skills) if resume.skills else resume.tech_keywords[:10]}
Resume Experience: {resume.experience_years} years
Resume Education: {', '.join(resume.education)}
Resume Full Text (first 1000 chars): {resume.full_text[:1000]}

Score the job (ats_score) by:
- keyword_overlap: 40% weight (keywords from resume found in JD)
- skills_match: 30% weight (resume skills match JD requirements)
- experience_align: 20% weight (experience level matches)
- education_match: 10% weight (education requirements met)
Provide the breakdown in ats_breakdown field.
"""
    
    try:
        response = await ai_client.complete(prompt)
        data = json.loads(response)
        
        # Determine match_verdict
        ats_score = data.get("ats_score", 50)
        if ats_score >= 80:
            match_verdict = "Perfect match"
        elif ats_score >= 55:
            match_verdict = "Should apply"
        else:
            match_verdict = "Not suitable"
        
        return JobResult(
            job_title=data.get("job_title", job.title),
            company=job.company,
            location=data.get("location", "Unknown"),
            remote_type=data.get("remote_type", "Unknown"),
            employment_type=data.get("employment_type", "Unknown"),
            experience_level=data.get("experience_level", "Unknown"),
            industry=data.get("industry", "Unknown"),
            company_size=data.get("company_size", "Unknown"),
            salary_min=data.get("salary_min"),
            salary_max=data.get("salary_max"),
            salary_currency=data.get("salary_currency"),
            salary_display=data.get("salary_display", "Not disclosed"),
            visa_sponsorship=data.get("visa_sponsorship", "Unknown"),
            posted_date=data.get("posted_date"),
            application_deadline=data.get("application_deadline"),
            description_summary=data.get("description_summary", ""),
            skills_required=data.get("skills_required", []),
            responsibilities=data.get("responsibilities", []),
            nice_to_have=data.get("nice_to_have", []),
            ats_score=ats_score,
            ats_breakdown=data.get("ats_breakdown", {}),
            match_verdict=match_verdict,
            keywords_matched=data.get("keywords_matched", []),
            keywords_missing=data.get("keywords_missing", []),
            skill_gaps=data.get("skill_gaps", []),
            resume_tips=data.get("resume_tips", []),
            cover_letter_hint=data.get("cover_letter_hint", ""),
            interview_topics=data.get("interview_topics", []),
            contact_email=data.get("contact_email"),
            contact_phone=data.get("contact_phone"),
            apply_url=job.url,
            source_board=job.source_board,
            ai_provider_used=provider_name,
            run_timestamp=datetime.now().isoformat()
        )
    except (json.JSONDecodeError, KeyError) as e:
        logger.warning(f"Failed to parse AI response for {job.title} at {job.company}: {e}")
        return None
    except Exception as e:
        logger.error(f"Error analyzing job {job.title}: {e}")
        return None

# ============================================================================
# EXCEL EXPORT
# ============================================================================

def create_excel_output(jobs: List[JobResult], output_dir: str, resume_path: str, search_role: str, 
                       search_location: str, provider_name: str):
    """Create richly formatted Excel file with 3 sheets"""
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_file = output_path / f"job_results_{timestamp}.xlsx"
    
    # Sort by ATS score descending
    jobs_sorted = sorted(jobs, key=lambda x: x.ats_score, reverse=True)
    
    wb = Workbook()
    
    # ========== SHEET 1: ALL JOBS ==========
    ws_all = wb.active
    ws_all.title = "All Jobs"
    
    # Headers
    headers = [
        "#", "match_verdict", "ats_score", "job_title", "company", "location", "remote_type",
        "salary_display", "employment_type", "experience_level", "industry", "company_size",
        "visa_sponsorship", "skills_required", "keywords_matched", "keywords_missing",
        "skill_gaps", "resume_tips", "cover_letter_hint", "interview_topics",
        "description_summary", "responsibilities", "nice_to_have",
        "contact_email", "contact_phone", "apply_url",
        "posted_date", "application_deadline", "source_board", "ai_provider_used", "run_timestamp"
    ]
    
    ws_all.append(headers)
    
    # Format header row
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    for cell in ws_all[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws_all.row_dimensions[1].height = 22
    ws_all.freeze_panes = "A2"
    
    # Data rows with colors based on match_verdict
    color_map = {
        "Perfect match": ("D6F0D6", "1A5C1A"),  # light green
        "Should apply": ("FFF4CC", "7A5C00"),   # light yellow
        "Not suitable": ("FFE5E5", "7A1A1A")    # light red
    }
    
    for idx, job in enumerate(jobs_sorted, start=2):
        row_data = [
            idx - 1,
            job.match_verdict,
            job.ats_score,
            job.job_title,
            job.company,
            job.location,
            job.remote_type,
            job.salary_display,
            job.employment_type,
            job.experience_level,
            job.industry,
            job.company_size,
            job.visa_sponsorship,
            "; ".join(job.skills_required),
            "; ".join(job.keywords_matched),
            "; ".join(job.keywords_missing),
            "; ".join(job.skill_gaps),
            "; ".join(job.resume_tips),
            job.cover_letter_hint,
            "; ".join(job.interview_topics),
            job.description_summary,
            "; ".join(job.responsibilities),
            "; ".join(job.nice_to_have),
            job.contact_email or "",
            job.contact_phone or "",
            job.apply_url,
            job.posted_date or "",
            job.application_deadline or "",
            job.source_board,
            job.ai_provider_used,
            job.run_timestamp
        ]
        ws_all.append(row_data)
        
        # Apply row color
        verdict = job.match_verdict
        if verdict in color_map:
            bg_color, font_color = color_map[verdict]
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            font = Font(color=font_color)
            for cell in ws_all[idx]:
                cell.fill = fill
                cell.font = font
    
    # Adjust column widths and wrap text
    wrap_columns = ["resume_tips", "cover_letter_hint", "interview_topics", "description_summary"]
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        if header in wrap_columns:
            ws_all.column_dimensions[col_letter].width = 40
            for row in ws_all.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            ws_all.column_dimensions[col_letter].width = min(max(12, len(header) + 2), 60)
    
    # Add autofilter
    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs_sorted) + 1}"
    
    # ========== SHEET 2: PERFECT MATCHES ==========
    perfect_jobs = [j for j in jobs_sorted if j.ats_score >= 80]
    
    ws_perfect = wb.create_sheet("Perfect Matches")
    ws_perfect.sheet_properties.tabColor = "1A5C1A"
    
    # Summary header
    ws_perfect.merge_cells("A1:J1")
    ws_perfect["A1"] = "Perfect Job Matches"
    ws_perfect["A1"].font = Font(bold=True, size=14)
    
    ws_perfect.merge_cells("A2:J2")
    if perfect_jobs:
        avg_score = sum(j.ats_score for j in perfect_jobs) / len(perfect_jobs)
        ws_perfect["A2"] = f"Total: {len(perfect_jobs)} | Avg ATS: {avg_score:.0f} | Date: {datetime.now().strftime('%Y-%m-%d')}"
    else:
        ws_perfect["A2"] = f"Total: 0 | Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws_perfect["A2"].font = Font(italic=True, size=10)
    
    ws_perfect.append([])  # Blank row
    
    # Headers
    ws_perfect.append(headers)
    for cell in ws_perfect[4]:
        cell.fill = PatternFill(start_color="D6F0D6", end_color="D6F0D6", fill_type="solid")
        cell.font = Font(bold=True, color="1A5C1A")
    
    ws_perfect.freeze_panes = "A5"
    
    # Data
    for idx, job in enumerate(perfect_jobs, start=5):
        row_data = [
            idx - 4,
            job.match_verdict,
            job.ats_score,
            job.job_title,
            job.company,
            job.location,
            job.remote_type,
            job.salary_display,
            job.employment_type,
            job.experience_level,
            job.industry,
            job.company_size,
            job.visa_sponsorship,
            "; ".join(job.skills_required),
            "; ".join(job.keywords_matched),
            "; ".join(job.keywords_missing),
            "; ".join(job.skill_gaps),
            "; ".join(job.resume_tips),
            job.cover_letter_hint,
            "; ".join(job.interview_topics),
            job.description_summary,
            "; ".join(job.responsibilities),
            "; ".join(job.nice_to_have),
            job.contact_email or "",
            job.contact_phone or "",
            job.apply_url,
            job.posted_date or "",
            job.application_deadline or "",
            job.source_board,
            job.ai_provider_used,
            job.run_timestamp
        ]
        ws_perfect.append(row_data)
        
        for cell in ws_perfect[idx]:
            cell.fill = PatternFill(start_color="D6F0D6", end_color="D6F0D6", fill_type="solid")
            cell.font = Font(color="1A5C1A")
    
    # ========== SHEET 3: RUN SUMMARY ==========
    ws_summary = wb.create_sheet("Run Summary")
    ws_summary.sheet_properties.tabColor = "0C447C"
    
    summary_data = calculate_summary_stats(jobs_sorted, resume_path, search_role, search_location, provider_name)
    
    # Write summary with styled formatting
    row = 1
    ws_summary.merge_cells(f"A{row}:B{row}")
    ws_summary[f"A{row}"] = "Job Hunt Summary Report"
    ws_summary[f"A{row}"].font = Font(bold=True, size=14, color="0C447C")
    row += 1
    
    ws_summary.merge_cells(f"A{row}:B{row}")
    ws_summary[f"A{row}"] = "─" * 50
    row += 1
    
    sections = [
        ("Run Info", [
            ("Run date & time", summary_data["run_timestamp"]),
            ("Resume file", Path(resume_path).name),
            ("Target role", search_role),
            ("Target location", search_location),
            ("AI provider used", provider_name),
            ("Boards scraped", summary_data["boards_scraped"]),
        ]),
        ("Job Statistics", [
            ("Total jobs found", summary_data["total_jobs_found"]),
            ("After deduplication", summary_data["jobs_after_dedup"]),
            ("Perfect matches (≥80)", summary_data["perfect_matches"]),
            ("Should apply (55-79)", summary_data["should_apply"]),
            ("Not suitable (<55)", summary_data["not_suitable"]),
        ]),
        ("Top Insights", [
            ("Top match", summary_data["top_match"]),
            ("Highest salary", summary_data["highest_salary"]),
            ("Remote opportunities", summary_data["remote_opportunities"]),
        ]),
        ("Skill Analysis", [
            ("Most common skill gap", summary_data["most_common_skill_gap"]),
            ("Most required skill", summary_data["most_required_skill"]),
        ])
    ]
    
    for section_name, items in sections:
        ws_summary.merge_cells(f"A{row}:B{row}")
        ws_summary[f"A{row}"] = section_name
        ws_summary[f"A{row}"].font = Font(bold=True, size=11, color="0C447C")
        row += 1
        
        for label, value in items:
            ws_summary[f"A{row}"] = label
            ws_summary[f"A{row}"].font = Font(bold=True)
            ws_summary[f"B{row}"] = str(value)
            row += 1
        
        row += 1
    
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 50
    
    # Save workbook
    wb.save(xlsx_file)
    logger.info(f"Excel file created: {xlsx_file}")
    
    return xlsx_file, summary_data

def calculate_summary_stats(jobs: List[JobResult], resume_path: str, role: str, location: str, 
                           provider: str) -> Dict[str, Any]:
    """Calculate summary statistics"""
    perfect = [j for j in jobs if j.ats_score >= 80]
    should_apply = [j for j in jobs if 55 <= j.ats_score < 80]
    not_suitable = [j for j in jobs if j.ats_score < 55]
    remote = [j for j in jobs if j.remote_type == "Remote"]
    
    # Top match
    top_match = jobs[0] if jobs else None
    top_match_str = f"{top_match.job_title} at {top_match.company} (ATS: {top_match.ats_score})" if top_match else "N/A"
    
    # Highest salary
    salaries = [j.salary_max for j in jobs if j.salary_max]
    highest_salary = max(salaries) if salaries else None
    highest_salary_str = f"{highest_salary} {jobs[0].salary_currency or ''}" if highest_salary else "Not disclosed"
    
    # Most common skill gap
    all_gaps = []
    for j in jobs:
        all_gaps.extend(j.skill_gaps)
    most_common_gap = Counter(all_gaps).most_common(1)[0][0] if all_gaps else "N/A"
    
    # Most required skill
    all_required = []
    for j in jobs:
        all_required.extend(j.skills_required)
    most_required_skill = Counter(all_required).most_common(1)[0][0] if all_required else "N/A"
    
    boards = list(set(j.source_board for j in jobs))
    
    return {
        "run_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_jobs_found": len(jobs),
        "jobs_after_dedup": len(jobs),  # Already deduped
        "perfect_matches": len(perfect),
        "should_apply": len(should_apply),
        "not_suitable": len(not_suitable),
        "top_match": top_match_str,
        "highest_salary": highest_salary_str,
        "remote_opportunities": len(remote),
        "most_common_skill_gap": most_common_gap,
        "most_required_skill": most_required_skill,
        "boards_scraped": ", ".join(boards) or "None"
    }

# ============================================================================
# EMAIL NOTIFICATION (OPTIONAL)
# ============================================================================

async def send_email_notification(jobs: List[JobResult], notify_email: str, search_role: str):
    """Send email with top matches if SendGrid is configured"""
    if not HAS_SENDGRID:
        logger.info("SendGrid not installed, skipping email notification")
        return
    
    sendgrid_key = os.getenv("SENDGRID_API_KEY", "").strip()
    if not sendgrid_key:
        logger.info("SENDGRID_API_KEY not set, skipping email notification")
        return
    
    try:
        perfect_jobs = [j for j in jobs if j.ats_score >= 80]
        top_10 = sorted(jobs, key=lambda x: x.ats_score, reverse=True)[:10]
        
        # Build HTML table
        html_rows = ""
        color_map = {
            "Perfect match": "#D6F0D6",
            "Should apply": "#FFF4CC",
            "Not suitable": "#FFE5E5"
        }
        
        for job in top_10:
            bg_color = color_map.get(job.match_verdict, "#FFFFFF")
            html_rows += f"""
            <tr style="background-color: {bg_color};">
                <td>{job.job_title}</td>
                <td>{job.company}</td>
                <td>{job.location}</td>
                <td>{job.ats_score}</td>
                <td>{job.match_verdict}</td>
                <td>{job.salary_display}</td>
                <td><a href="{job.apply_url}">Apply</a></td>
            </tr>
            """
        
        html_content = f"""
        <html>
            <body>
                <h2>Job Hunt Results - {search_role}</h2>
                <p><strong>Perfect Matches:</strong> {len(perfect_jobs)}</p>
                <p><strong>Run Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                
                <h3>Top 10 Opportunities</h3>
                <table border="1" cellpadding="10" style="border-collapse: collapse;">
                    <tr style="background-color: #2D2D2D; color: white;">
                        <th>Job Title</th>
                        <th>Company</th>
                        <th>Location</th>
                        <th>ATS Score</th>
                        <th>Match Verdict</th>
                        <th>Salary</th>
                        <th>Action</th>
                    </tr>
                    {html_rows}
                </table>
            </body>
        </html>
        """
        
        sg = SendGridAPIClient(sendgrid_key)
        email = Mail(
            from_email="noreply@jobhunter.local",
            to_emails=notify_email,
            subject=f"Job Hunt Results — {len(perfect_jobs)} Perfect Matches ({search_role})",
            html_content=html_content
        )
        
        response = sg.send(email)
        if response.status_code == 202:
            logger.info(f"Email sent to {notify_email}")
        else:
            logger.warning(f"Email send failed with status {response.status_code}")
    except Exception as e:
        logger.warning(f"Failed to send email: {e}")

# ============================================================================
# SUMMARY JSON OUTPUT
# ============================================================================

def write_summary_json(summary_data: Dict[str, Any], output_dir: str):
    """Write summary.json for GitHub Step Summary"""
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    json_file = output_path / "summary.json"
    with open(json_file, "w") as f:
        json.dump(summary_data, f, indent=2)
    
    logger.info(f"Summary JSON written to {json_file}")

# ============================================================================
# MAIN ORCHESTRATION
# ============================================================================

async def main():
    """Main orchestration"""
    global logger
    
    # Parse CLI arguments
    parser = argparse.ArgumentParser(description="Job Hunter - AI-powered job hunting agent")
    parser.add_argument("--resume", required=True, help="Path to resume (PDF or DOCX)")
    parser.add_argument("--role", required=True, help="Target job role")
    parser.add_argument("--location", default="Remote", help="Target location")
    parser.add_argument("--limit", type=int, default=100, help="Max jobs to process")
    parser.add_argument("--ats-threshold", type=int, default=0, help="Min ATS score to include")
    parser.add_argument("--boards", default="linkedin,indeed,glassdoor,naukri,wellfound", 
                       help="Comma-separated list of boards")
    parser.add_argument("--output-dir", default="./output", help="Output directory")
    parser.add_argument("--provider", help="Force specific AI provider")
    parser.add_argument("--log-level", default="INFO", help="Log level (DEBUG|INFO|WARNING)")
    
    args = parser.parse_args()
    
    # Setup logging
    logger = setup_logging(args.log_level)
    logger.info("=" * 70)
    logger.info("Job Hunter Started")
    logger.info("=" * 70)
    
    # Load environment
    load_dotenv()
    
    # Detect AI provider
    logger.info("Detecting AI provider...")
    ai_client, active_provider = await detect_active_provider(args.provider)
    
    try:
        async with ai_client:
            # Parse resume
            logger.info(f"Parsing resume: {args.resume}")
            resume_text = parse_resume_file(args.resume)
            logger.info(f"Resume loaded: {len(resume_text)} characters")
            
            # Extract resume profile using AI
            logger.info("Extracting resume profile with AI...")
            resume = await extract_resume_profile(resume_text, ai_client)
            logger.info(f"Resume profile extracted: {resume.experience_years} years exp, {len(resume.tech_keywords)} tech keywords")
            
            # Scrape jobs
            boards = [b.strip().lower() for b in args.boards.split(",")]
            logger.info(f"Scraping jobs from boards: {boards}")
            raw_jobs = await scrape_all_jobs(args.role, args.location, boards, args.limit)
            logger.info(f"Jobs found: {len(raw_jobs)}")
            
            if not raw_jobs:
                logger.warning("No jobs found. Exiting.")
                return
            
            # Analyze jobs with progress bar
            logger.info("Analyzing jobs with AI...")
            analyzed_jobs = []
            
            with tqdm(total=len(raw_jobs), desc="Analyzing jobs") as pbar:
                for raw_job in raw_jobs:
                    result = await analyze_job(raw_job, resume, ai_client, active_provider)
                    if result:
                        analyzed_jobs.append(result)
                    pbar.update(1)
            
            # Filter by ATS threshold
            if args.ats_threshold > 0:
                analyzed_jobs = [j for j in analyzed_jobs if j.ats_score >= args.ats_threshold]
                logger.info(f"After ATS threshold filter: {len(analyzed_jobs)} jobs")
            
            # Create Excel output
            logger.info("Creating Excel output...")
            xlsx_file, summary_data = create_excel_output(analyzed_jobs, args.output_dir, 
                                                          args.resume, args.role, args.location, 
                                                          active_provider)
            
            # Write summary JSON
            write_summary_json(summary_data, args.output_dir)
            
            # Send email if configured
            notify_email = os.getenv("NOTIFY_EMAIL", "").strip()
            if notify_email:
                logger.info(f"Sending notification email to {notify_email}...")
                await send_email_notification(analyzed_jobs, notify_email, args.role)
            
            logger.info("=" * 70)
            logger.info("Job Hunter Completed Successfully")
            logger.info(f"Output: {xlsx_file}")
            logger.info(f"Perfect Matches: {summary_data['perfect_matches']}")
            logger.info(f"Should Apply: {summary_data['should_apply']}")
            logger.info("=" * 70)
    
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)

if __name__ == "__main__":
    asyncio.run(main())
